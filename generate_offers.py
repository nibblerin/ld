import os
import json
from dotenv import load_dotenv
from groq import Groq
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

load_dotenv()

client = Groq(
    api_key=os.getenv("GROQ_API_KEY")
)

SELECTED_FILE = "candidates_search_results/selected_candidates.json"
BLOGGERS_FILE = "json_candidates/clean_candidates_for_llm_shortened.json"
EXCEL_FILE = "candidates_search_results/selected_candidates.xlsx"

with open(
    SELECTED_FILE,
    encoding="utf-8"
) as f:
    selected = json.load(f)

with open(
    BLOGGERS_FILE,
    encoding="utf-8"
) as f:
    bloggers = json.load(f)

first_selected = selected["selected"][1]

selected_username = first_selected["username"]

blogger = None

for item in bloggers:
    username = (
        item
        .get("profile", {})
        .get("username")
    )
    if username == selected_username:
        blogger = item
        break

if not blogger:
    raise Exception(
        "Blogger not found"
    )

profile = blogger.get(
    "profile",
    {}
)

content = blogger.get(
    "content",
    {}
)

metrics = blogger.get(
    "metrics",
    {}
)

visual = blogger.get(
    "visual_style",
    []
)

blogger_data = json.dumps(
    {
        "profile": profile,
        "posts": content.get(
            "posts",
            []
        )[:10],
        "metrics": metrics,
        "visual_style": visual
    },
    ensure_ascii=False
)


completion = client.chat.completions.create(
    model="openai/gpt-oss-120b",
    messages=[
        {
            "role":"user",
            "content":prompt
        }
    ],
    temperature=0.5,
    max_completion_tokens=1000,

)
offer = completion.choices[0].message.content.strip()

if os.path.exists(EXCEL_FILE):
    wb = load_workbook(EXCEL_FILE)
else:
    import pandas as pd

    df = pd.read_csv(
        "candidates_search_results/selected_candidates.csv",
        encoding="utf-8-sig"
    )
    df.to_excel(
        EXCEL_FILE,
        index=False
    )

    wb = load_workbook(EXCEL_FILE)

ws = wb.active

headers = {}

for cell in ws[1]:
    headers[cell.value] = cell.column

if "offer" not in headers:
    offer_column = ws.max_column + 1
    ws.cell(
        row=1,
        column=offer_column,
        value="offer"
    )
else:
    offer_column = headers["offer"]

username_column = headers["Username"]

for row in range(
    2,
    ws.max_row + 1
):

    username = ws.cell(
        row=row,
        column=username_column
    ).value

    if username == selected_username:

        ws.cell(
            row=row,
            column=offer_column,
            value=offer
        )

        ws.row_dimensions[row].height = 180

        break

for row in ws.iter_rows():

    for cell in row:

        cell.alignment = Alignment(
            wrap_text=True,
            vertical="top"
        )

        if cell.row == 1:
            cell.font = Font(
                bold=True
            )

ws.column_dimensions["A"].width = 35
ws.column_dimensions["B"].width = 25
ws.column_dimensions["C"].width = 70
ws.column_dimensions["D"].width = 90

ws.freeze_panes = "A2"

wb.save(
    EXCEL_FILE
)

print(
    "Offer added for:",
    selected_username
)